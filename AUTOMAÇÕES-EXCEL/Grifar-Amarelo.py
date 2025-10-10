import customtkinter as ctk
from tkinter import filedialog
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

def escolher_sintese():
    global sintese
    sintese = filedialog.askopenfilename(
        filetypes = [("Excel files", "*.xlsx")],
        title = "Selecione o arquivo SINTESE"
    )
    print("Arquivo SINTESE: ", sintese)
    if sintese:
        btn1.configure(text = "✅")
    return sintese

def escolher_art():
    global art
    art = filedialog.askopenfilename(
        filetypes = [("CSV files", "*.csv")],
        title = "Selecione o arquivo ART"
    )
    print("Arquivo ART: ", art)
    if art:
        btn2.configure(text = "✅")
    return art

# Função que faz a leitura e grifa
def processar():
    global msg_label_error
    global msg_label
    
    msg = ""
    if not sintese:
        msg += "SÍNTESE NAO SELECIONADA❗\n"
    if not art:
        msg += "ART NAO SELECIONADO❗"
    elif sintese and art:
        msg_label_error.destroy()
    if msg != "":
        try:
            msg_label_error
        except NameError:
            msg_label_error = ctk.CTkLabel(root, text = "", font = ("Arial", 16))
            msg_label_error.pack(padx = 20, pady = 15)
        msg_label_error.configure(text = msg)
    else:
        global celulas_ajuste
        celulas_ajuste = []

        dfSintese = pd.read_excel(sintese)
        dfArt = pd.read_csv(art, encoding = "latin1", sep = ";")

        dfSintese.rename(columns = {20: "MAPA"}, inplace = True)

        infoArt = dfArt["Mapa"].tolist()
        infoSintese = dfSintese["MAPA"].tolist()

        # Abre o Excel original com openpyxl
        wb = load_workbook(sintese)
        ws = wb.active

        fill_amarelo = PatternFill(start_color = "FFFF00", end_color = "FFFF00", fill_type = "solid")

        # Percorre e grifa se encontrar correspondência, ao contrário, guarda em uma lista a posição e o Valor da Célula
        for i, mapa_sintese in enumerate(infoSintese, start=2):  # Excel começa em 1 + header
            #Verifica se eh "nan" ( aparece quando nao tem nada digitado na célula do Excel )
            if not pd.isna(mapa_sintese) and mapa_sintese != "MAPA":
                cell = ws.cell(row=i, column=1)
                if mapa_sintese in infoArt:
                    cell.fill = fill_amarelo
                else:
                    celulas_ajuste.append(f"- N{i}: {mapa_sintese}")

        btn1.configure(text = "SELECIONAR SÍNTESE")
        btn2.configure(text = "SELECIONAR ART")
        try:
            msg_label
        except:
            msg_label_save = ctk.CTkLabel(root, text = "", font = ("Arial", 16))
            msg_label_save.pack(padx = 20, pady = 15)
            
            msg_label = ctk.CTkLabel(root, text = "", font = ("Arial", 16))
            msg_label.pack(padx = 20, pady = 15)

        # Salva em um outro arquivo
        novo_arquivo = sintese.replace(".xlsx", "_COLORIDO.xlsx")
        wb.save(novo_arquivo)
        msg_label_save.configure(text = "Arquivo salvo com sucesso ✅")
        print("Arquivo salvo com sucesso ✅")

        # Monta lista de mapas que faltaram
        if len(celulas_ajuste) != 0:
            texto = "\n".join(celulas_ajuste)
            msg_label.configure(text = f"Faltam os seguintes Mapas: \n{texto}")
        else:
            msg_label.configure(text = f"Todos os Mapas estão Preenchidos CORRETAMENTE ✅")

sintese = None
art = None

# GUI com CustomTkinter
ctk.set_appearance_mode("dark")
ctk.set_default_color_theme("green")

root = ctk.CTk()
root.title("Processador de Arquivos")

btn1 = ctk.CTkButton(root, text = "SELECIONAR SINTESE", command = escolher_sintese, width = 300, height = 50, font = ("Arial", 20, "bold"))
btn1.pack(padx = 20, pady = 15)

btn2 = ctk.CTkButton(root, text = "SELECIONAR ART", command = escolher_art, width = 300, height = 50, font = ("Arial", 20, "bold"))
btn2.pack(padx = 20, pady = 15)

btn3 = ctk.CTkButton(root, text = "PROCESSAR", command = processar, width = 300, height = 50, font = ("Arial", 20, "bold"))
btn3.pack(padx = 20, pady = 25)

root.mainloop()